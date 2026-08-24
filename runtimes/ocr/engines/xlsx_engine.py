"""Native XLSX parser (v1.20.0 Phase 2). openpyxl, not Docling, no OCR."""

from __future__ import annotations

import io
from typing import Optional

from ..documents import MAX_PAGES, DocumentError, decode_payload
from .base import ParsedDocument, ParsedPage, ProgressFn
from .office_common import markdown_table, raise_office_open_error

MAX_SHEET_ROWS = 200


class XlsxEngine:
    name = "xlsx"

    def parse(
        self,
        document_base64: str,
        *,
        dpi: Optional[int] = None,
        max_pages: int = MAX_PAGES,
        progress: Optional[ProgressFn] = None,
    ) -> ParsedDocument:
        del dpi
        data = decode_payload(document_base64)
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise DocumentError(
                "engine-unavailable",
                f"openpyxl is not installed: {exc}",
            ) from exc

        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise_office_open_error("Excel", exc)

        try:
            sheets = list(workbook.worksheets)
            capped = max(1, min(max_pages, MAX_PAGES, len(sheets) or 1))
            chosen = sheets[:capped]
            if not chosen:
                return ParsedDocument(
                    pages=(ParsedPage(index=0, text=""),),
                    engine=self.name,
                    markdown=None,
                )

            parsed: list[ParsedPage] = []
            for index, sheet in enumerate(chosen):
                text = _render_sheet(sheet)
                parsed.append(ParsedPage(index=index, text=text))
                if progress is not None:
                    progress(index + 1, len(chosen))

            joined = "\n\n".join(page.text for page in parsed).strip()
            return ParsedDocument(
                pages=tuple(parsed),
                engine=self.name,
                markdown=joined or None,
            )
        except DocumentError:
            raise
        except Exception as exc:
            raise_office_open_error("Excel", exc)
        finally:
            workbook.close()


def _render_sheet(sheet: object) -> str:
    title = getattr(sheet, "title", None) or "Sheet"
    rows: list[list[str]] = []
    truncated = False
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        if index >= MAX_SHEET_ROWS:
            truncated = True
            break
        rows.append(["" if cell is None else str(cell) for cell in row])
    heading = f"# {title}"
    table = markdown_table(rows) if rows else ""
    parts = [heading]
    if table:
        parts.append(table)
    if truncated:
        parts.append(f"_Truncated to {MAX_SHEET_ROWS} rows._")
    return "\n\n".join(parts)
